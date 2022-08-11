#Just python not related to selenium
import openpyxl

workbook = openpyxl.load_workbook("DataFile.xlsx") #stored in an object workbook i.e, workbook is an object of "Workbook" class
#print(type(workbook))

sheets = workbook.sheetnames
#print(sheets) #returns a list of sheets present
#print(workbook.active.title) #current active/open sheet
#print(sheets[1])

'''Access record from a particular cell'''

'''option 1'''
sheet1 = workbook['Names'] #get the sheet object
print(sheet1["B2"].value)

'''option 2'''
print(workbook["Names"]["A1"].value)

'''option 3'''
print(sheet1.cell(3,2).value) #3-row,2-column ...not index
print(sheet1.cell(3,3).value)


sheet2 = workbook['Marks']
print(sheet2.cell(2,2).value)
print(sheet2.cell(3,1).value)
c = sheet2.cell(1,2)
#print(type(c)) # c is an object of "Cell" class

'''option 4'''
print(sheet2.cell(row=3,column=2).value)


#to print all data

rows = sheet1.max_row
columns = sheet1.max_column
print(rows)
print(columns)

for i in range(1,rows+1):   #outer loop for rows, range wont consider the last row, therfore row+1
    for j in range(1,columns+1):       #inner loop for columns, range wont consider the last col, therfore col+1
       print(sheet1.cell(i,j).value)


'''add a row to the sheet and save it by creating a new excel'''
sheet2.cell(row=5,column=1,value="Wilson")
sheet2.cell(row=5,column=2,value="20")
workbook.save("Report.xlsx") #save and create a new excel

'''or'''
sheet2['A6'].value = "Martha"
sheet2['B6'].value = "100"
workbook.save("DataFile.xlsx") #saving in same file