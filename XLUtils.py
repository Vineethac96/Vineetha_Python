import  openpyxl

def getRowCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return(sheet.max_row)

def getColCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return(sheet.max_column)

def readData(file, sheetName,rowno,colno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(row=rowno,column=colno).value

def writeData(file, sheetName,rowno,colno,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    data = sheet.cell(row=rowno,column=colno).value
    workbook.save(file)