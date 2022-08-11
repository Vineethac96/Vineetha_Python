#Just python not related to selenium

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill

wb = Workbook() #creating "wb" object from Workbook class

sheet1 = wb.active #default, the first sheet will be open and active

sheet1['A1'].value = "Name"
sheet1['B1'].value = "Age"
sheet1['C1'].value = "Status"
sheet1['A2'].value = "Vineetha"
sheet1['B2'].value = "26"
sheet1['C2'].value = "Active"
sheet1['C2'].fill = PatternFill("solid",fgColor='33FF39')
sheet1['A3'].value = "kkkk"
sheet1['B3'].value = "120"
sheet1['C3'].value = "Inactive"
sheet1['C3'].fill = PatternFill("solid",fgColor='FF6E33')
sheet1['A4'].value = "Manoj"
sheet1['B4'].value = "29"
sheet1['C4'].value = "Active"
sheet1['C4'].fill = PatternFill("solid",fgColor='33FF39')

wb.save("C:/Users/sumanth/Desktop/Demo.xlsx")


